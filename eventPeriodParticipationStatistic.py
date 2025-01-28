import os

# Установка необходимых библиотек
os.system("pip install openpyxl")

input("Для запуска программы нажмите Enter")

from openpyxl import Workbook, load_workbook

# Получаем список всех файлов .csv в текущей папке
csv_files = [f for f in os.listdir() if f.endswith('.csv')]

for file in csv_files:
    # Загружаем файл
    wb = Workbook()
    ws = wb.active

    # Читаем данные из файла
    with open(file, 'r', encoding='utf-8') as f:
        for row in f:
            # Заменяем ';' на ',' для разделения ячеек
            ws.append(row.strip().split(';'))

    # Удаляем дубликаты в столбце I
    seen = set()
    rows_to_keep = []

    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):  # max_col=9 для столбца I
        if row[8] not in seen:  # Индекс 8 соответствует столбцу I (0-индексация)
            seen.add(row[8])
            rows_to_keep.append(row)

    # Очищаем лист и записываем уникальные строки
    ws.delete_rows(2, ws.max_row)  # Удаляем все строки, начиная со 2-й
    for row in rows_to_keep:
        ws.append(row)

    # Сохраняем изменения в том же файле
    wb.save(file)

print("Дубликаты удалены из всех файлов .csv в текущей папке.")

# Закрыть браузер
input("Нажмите Enter для закрытия окна")

# softy_plug